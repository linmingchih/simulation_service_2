import os
import shutil
import zipfile
import openpyxl
from pyedb import Edb
from app.utils import is_file_locked

def apply_xlsx(xlsx_path, edb_path, edb_version="2024.1"):
    edb = Edb(edb_path, edbversion=edb_version)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Stackup"]

    unit = str(ws.cell(row=1, column=2).value or "mm").lower()

    material_dic = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        layer_name, layer_type, thickness_val, permittivity, loss_tangent, conductivity = row
        if unit == "mil":
            thickness_m = float(thickness_val) * 25.4 / 1000.0
        else:
            thickness_m = float(thickness_val) / 1000.0
        edb.stackup.stackup_layers[layer_name].thickness = thickness_m
    
        if layer_type == "signal":
            if conductivity not in material_dic:
                name = f'metal_{conductivity}'
                mat = edb.materials.add_conductor_material(name, conductivity)
                material_dic[conductivity] = mat
                
            edb.stackup.stackup_layers[layer_name].material = material_dic[conductivity].name
        else:
            if (permittivity, loss_tangent) not in material_dic:
                name = f'dielectric_{permittivity}_{loss_tangent}'
                mat = edb.materials.add_dielectric_material(name, 
                                                            permittivity, 
                                                            loss_tangent)
                material_dic[(permittivity, loss_tangent)] = mat
            
            edb.stackup.stackup_layers[layer_name].material = material_dic[(permittivity, loss_tangent)].name
    
    edb.save()
    edb.close_edb()


def run(job_path, data=None, files=None, config=None):
    input_dir = os.path.join(job_path, 'input')
    output_dir = os.path.join(job_path, 'output')
    os.makedirs(input_dir, exist_ok=True)
    os.makedirs(output_dir, exist_ok=True)

    edb_version = (config or {}).get("edb_version", "2024.1")
    xlsx = files.get('xlsx_file') if files else None
    if xlsx and xlsx.filename:
        x_path = os.path.join(input_dir, xlsx.filename)
        xlsx.save(x_path)

        if is_file_locked(x_path) or not os.access(x_path, os.W_OK):
            os.remove(x_path)
            return {
                'error': (
                    '\u26a0\ufe0f The selected Excel file is currently locked or '
                    'in use. Please close it and try again.'
                )
            }

        shutil.copy(x_path, os.path.join(output_dir, 'updated.xlsx'))

        edb_dir = os.path.join(output_dir, 'design.aedb')
        if os.path.isdir(edb_dir):
            apply_xlsx(x_path, edb_dir, edb_version)

            # Generate layer images for visualization
            try:
                edb = Edb(edb_dir, edbversion=edb_version)
                for layer_name in edb.stackup.signal_layers:
                    img_path = os.path.join(output_dir, f"{layer_name}.png")
                    edb.nets.plot(layers=layer_name, save_plot=img_path)
                edb.close()
            except Exception:
                pass

            # Create a zipped copy of the updated AEDB for easy download
            zip_path = os.path.join(output_dir, 'updated_pyedb.zip')
            if os.path.isfile(zip_path):
                os.remove(zip_path)
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root, dirs, files in os.walk(edb_dir):
                    for file in files:
                        abs_path = os.path.join(root, file)
                        rel_path = os.path.relpath(abs_path, os.path.dirname(edb_dir))
                        zf.write(abs_path, rel_path)

    return {}
